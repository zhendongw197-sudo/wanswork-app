# -*- coding: utf-8 -*-
"""导出模块：工单 Excel 导出与汇总统计。"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# 表头与行字典字段的对应关系
HEADERS = [
    '流程编号', '流程类型', '户号', '户名', '客户经理',
    '流程开始日期', '上门服务截止日', '全流程截止日', '剩余时间',
    '地址', '是否建群', '状态',
]
FIELDS = [
    'flow_id', 'flow_type', 'account_no', 'account_name', 'manager',
    'start_time', 'visit_deadline', 'full_deadline', 'remaining',
    'address', 'in_group', 'status',
]

# 各列宽度（字符数）
COLUMN_WIDTHS = [18, 12, 16, 16, 12, 20, 20, 20, 14, 30, 8, 8]


def export_orders(path: str, rows: list) -> str:
    """把工单行导出为 Excel 文件，返回文件路径。

    rows 元素为 dict，键见 FIELDS；缺省字段导出为空字符串。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = '工单列表'

    # 表头样式
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4472C4')
    header_align = Alignment(horizontal='center', vertical='center')

    ws.append(HEADERS)
    for col_idx, width in enumerate(COLUMN_WIDTHS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 数据行
    for row in rows:
        ws.append([_cell_text(row.get(f, '')) for f in FIELDS])

    ws.freeze_panes = 'A2'  # 冻结表头
    wb.save(path)
    return path


def _cell_text(value) -> str:
    """单元格取值：None 转为空串，其余转字符串。"""
    if value is None:
        return ''
    return str(value)


# ---------- 供方单台账导出 ----------

# 17 个材料键与表头（与 app/supply.py 的 MATERIAL_COLUMNS 保持一致，此处写字面量以保持本模块独立）
SUPPLY_MATERIAL_FIELDS = [
    'cable_2x16', 'cable_4x16', 'cable_4x35', 'cable_4x70', 'cable_4x150', 'cable_4x240',
    'box_wall', 'box_floor', 'mpp100', 'tray_200x100', 'tray_300x150',
    'term_16', 'term_35', 'term_70', 'term_150', 'term_240', 'overhead',
]
SUPPLY_MATERIAL_HEADERS = [
    '2*16', '4*16', '4*35', '4*70', '4*150', '4*240',
    '分支箱（挂壁）', '分支箱（落地）', 'MPP100管子',
    '桥架（200*100）', '桥架（300*150）',
    '16冷缩终端', '35冷缩终端', '70冷缩终端', '150冷缩终端', '240冷缩终端', '架空导线',
]

# 供方单台账列表头（14 个基本字段 + 17 个材料，共 31 列）
SUPPLY_HEADERS = [
    '客户经理', '是否归档', '户号', '流程编号', '户名', '地址',
    '流程开始时间', '方案推送时间', '流程超期时间',
    '当前进度（营销部反馈）', '设备部进度（是否发料）', '备注', '配套单位',
    '现场勘察方案及工作量',
] + SUPPLY_MATERIAL_HEADERS

# 表头对应的字段键；材料键从 record['materials'] 取值，其余取顶层键
SUPPLY_FIELDS = [
    'manager', 'archived', 'account_no', 'flow_id', 'account_name', 'address',
    'start_time', 'push_time', 'overdue_time',
    'progress_marketing', 'progress_equipment', 'remark', 'contractor',
    'work_desc',
] + SUPPLY_MATERIAL_FIELDS

# 供方单台账各列宽度（字符数）
SUPPLY_COLUMN_WIDTHS = (
    [10, 10, 20, 18, 18, 30, 20, 20, 20, 24, 24, 24, 12, 40]
    + [14] * 17
)


def export_supply_orders(path: str, records: list) -> str:
    """把供方单台账记录导出为 Excel 文件，返回文件路径。

    records 元素为 dict：基本字段取顶层键，材料字段取 record['materials']；
    缺省字段导出为空字符串。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = '供方单台账'

    # 表头样式（与工单列表一致）
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4472C4')
    header_align = Alignment(horizontal='center', vertical='center')

    ws.append(SUPPLY_HEADERS)
    for col_idx, width in enumerate(SUPPLY_COLUMN_WIDTHS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 数据行：前 14 列取顶层键，后 17 列材料从 materials 取
    base_count = len(SUPPLY_FIELDS) - len(SUPPLY_MATERIAL_FIELDS)
    for record in records:
        materials = record.get('materials') or {}
        row = []
        for idx, field in enumerate(SUPPLY_FIELDS):
            if idx < base_count:
                row.append(_cell_text(record.get(field, '')))
            else:
                row.append(_cell_text(materials.get(field, '')))
        ws.append(row)

    ws.freeze_panes = 'A2'  # 冻结表头
    wb.save(path)
    return path


def summary(rows: list) -> dict:
    """汇总统计：{'total': n, 'by_type': {类型: n}, 'by_status': {状态: n}}。"""
    by_type: dict = {}
    by_status: dict = {}
    for row in rows:
        t = row.get('flow_type', '') or '未填写'
        s = row.get('status', '') or '未填写'
        by_type[t] = by_type.get(t, 0) + 1
        by_status[s] = by_status.get(s, 0) + 1
    return {'total': len(rows), 'by_type': by_type, 'by_status': by_status}
