# -*- coding: utf-8 -*-
"""端到端集成冒烟测试（无界面、纯后台链路）。

链路：临时目录建 Store → 录 3 条不同类型工单（含周末 / 17:30 后边界用例）
→ 验证截止日计算 → 生成终止材料与四方交底全套文件 → 导出 Excel → 汇总统计。

用法（项目根目录或任意目录下）：
    python tests\\test_e2e.py
"""

import os
import shutil
import sys
import tempfile
from datetime import date, datetime

# 保证可以 import app 包（项目根目录加入 sys.path）
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_ROOT)

from app import export, gen_call, gen_docs, gen_sms, worktime  # noqa: E402
from app.store import Store  # noqa: E402

# 基准日历：2024-01-05 周五，01-06 周六，01-07 周日，01-08 周一

ORDERS = [
    # 边界用例1：周六 10:00 报装 → 起算顺延到下周一 8:30
    {
        'flow_id': 'GD20240106001', 'flow_type': '低压居民新装',
        'account_no': '3201240001111111', 'account_name': '周末小吃店',
        'manager': '王五', 'status': '在办',
        'phone': '13800000001', 'address': '溧水区永阳街道周末路1号',
        'start_time': '2024-01-06 10:00:00',
    },
    # 边界用例2：周五 18:00（17:30 后）报装 → 起算顺延到下周一 8:30
    {
        'flow_id': 'GD20240105002', 'flow_type': '低压非居民新装',
        'account_no': '3201240002222222', 'account_name': '加班建材厂',
        'manager': '王五', 'status': '在办',
        'phone': '13800000002', 'address': '溧水区柘塘街道加班路2号',
        'start_time': '2024-01-05 18:00:00',
    },
    # 常规用例：周一 09:00 工作时间内报装 → 起算不变
    {
        'flow_id': 'GD20240108003', 'flow_type': '低压充电桩新装',
        'account_no': '3201240003333333', 'account_name': '张三',
        'manager': '王五', 'status': '在办',
        'phone': '13800000003', 'address': '溧水区永阳街道常规路3号',
        'start_time': '2024-01-08 09:00:00',
    },
]

# 预期截止日（手工推算，24 小时累计制）：
# 工单1：起算 周一01-08 08:30，full=5 天（120h）
#   = 15.5h(周一) + 24*4(周二~周五) + 8.5h(下周一) → 2024-01-15 08:30
EXPECTED_FULL_1 = datetime(2024, 1, 15, 8, 30)
# 工单2：起算 周一01-08 08:30，visit=4 天（96h）
#   = 15.5 + 24*3 + 8.5 → 2024-01-12 08:30；full=15 天（360h）
#   = 15.5 + 24*14 + 8.5 → 2024-01-29 08:30
EXPECTED_VISIT_2 = datetime(2024, 1, 12, 8, 30)
EXPECTED_FULL_2 = datetime(2024, 1, 29, 8, 30)
# 工单3：起算 周一01-08 09:00 不变，full=15 天（360h）
#   = 15h(周一) + 24*14 + 9h → 2024-01-29 09:00；visit=3 天（72h）
#   = 15h(周一) + 24*2(周二~周三) + 9h(周四) → 2024-01-11 09:00
EXPECTED_FULL_3 = datetime(2024, 1, 29, 9, 0)
EXPECTED_VISIT_3 = datetime(2024, 1, 11, 9, 0)

CONFIG_KEYS = [
    'gas_company', 'builder_name', 'builder_leader', 'builder_phone',
    'org_name', 'company_short', 'dig_depth_cm',
]


def _fmt_dt(dt):
    return '-' if dt is None else dt.strftime('%Y-%m-%d %H:%M:%S')


def main() -> int:
    tmp = tempfile.mkdtemp(prefix='workorder_e2e_')
    data_dir = os.path.join(tmp, 'data')
    out_dir = os.path.join(tmp, 'output')
    os.makedirs(out_dir, exist_ok=True)
    try:
        store = Store(data_dir)
        store.add_manager('王五', '13912345678')

        # ---- 1. 录入 3 条工单 ----
        for order in ORDERS:
            store.add_order(order)
        assert len(store.list_orders()) == 3, '工单数量应为 3'
        print('[1] 录入 3 条工单（含周末 / 17:30 后边界用例）：通过')

        # ---- 2. 验证截止日计算 ----
        holidays = store.get_holidays()
        extra = store.get_extra_workdays()

        def deadlines_of(order):
            start = datetime.strptime(order['start_time'], '%Y-%m-%d %H:%M:%S')
            return worktime.compute_deadlines(order['flow_type'], start, holidays, extra)

        v1, f1 = deadlines_of(ORDERS[0])
        assert v1 is None and f1 == EXPECTED_FULL_1, f'工单1截止日错误: {v1}, {f1}'
        v2, f2 = deadlines_of(ORDERS[1])
        assert v2 == EXPECTED_VISIT_2 and f2 == EXPECTED_FULL_2, f'工单2截止日错误: {v2}, {f2}'
        v3, f3 = deadlines_of(ORDERS[2])
        assert v3 == EXPECTED_VISIT_3 and f3 == EXPECTED_FULL_3, f'工单3截止日错误: {v3}, {f3}'
        print('[2] 截止日计算（周末顺延 / 17:30 后顺延 / 常规）：通过')

        # ---- 2b. 节假日保存后重算（模拟界面保存节假日 → 截止日联动变化）----
        store.set_holidays({date(2024, 1, 9)})  # 周二放假一天
        holidays = store.get_holidays()
        _v3b, f3b = deadlines_of(ORDERS[2])
        assert f3b == datetime(2024, 1, 30, 9, 0), f'节假日重算错误: {f3b}'
        store.set_holidays(set())  # 还原
        holidays = store.get_holidays()
        print('[2b] 节假日保存后截止日重算（周二放假 → 截止日顺延一天）：通过')

        # ---- 3. 生成终止材料全套（通话截图 + 短信截图 + 情况说明）----
        order = dict(ORDERS[0])
        order['manager_phone'] = store.get_manager_phone(order['manager'])
        term_dir = os.path.join(out_dir, f"终止材料_{order['flow_id']}")
        os.makedirs(term_dir, exist_ok=True)
        apply_date = '2024年1月6日'
        p1 = gen_call.gen_call_screenshot(
            order['phone'], os.path.join(term_dir, '材料1-通话记录截图.png'),
            duration_sec=10, battery=79)
        p2 = gen_sms.gen_sms_screenshot(
            customer_name=order['account_name'], manager_name=order['manager'],
            apply_date=apply_date, address=order['address'],
            flow_type=order['flow_type'], flow_id=order['flow_id'],
            reason='暂无用电需求', manager_phone=order['manager_phone'],
            out_path=os.path.join(term_dir, '材料2-短信确认截图.png'), battery=79)
        p3 = gen_docs.gen_termination_note(
            order, '暂无用电需求',
            os.path.join(term_dir, '材料4-终止白名单情况说明.docx'))
        for p in (p1, p2, p3):
            assert os.path.exists(p) and os.path.getsize(p) > 0, f'文件未生成: {p}'
        print(f'[3] 终止材料全套 3 个文件生成：通过（{term_dir}）')

        # ---- 4. 生成四方交底全套（交底申请书 + 施工方案 + 监理情况说明）----
        order2 = dict(ORDERS[1])
        order2['manager_phone'] = store.get_manager_phone(order2['manager'])
        config = {k: store.get_config(k, '') for k in CONFIG_KEYS}
        assert all(k in config for k in CONFIG_KEYS), 'config 七键不完整'
        try:
            config['dig_depth_cm'] = int(config['dig_depth_cm'])
        except (TypeError, ValueError):
            pass
        jd_dir = os.path.join(out_dir, f"四方交底_{order2['flow_id']}")
        os.makedirs(jd_dir, exist_ok=True)
        q1 = gen_docs.gen_jiaodi_apply(
            order2, '柘塘街道', '20', '2024年1月10日',
            os.path.join(jd_dir, '交底申请书.docx'), config)
        q2 = gen_docs.gen_construct_plan(
            order2, '20', os.path.join(jd_dir, '施工方案.docx'), config)
        q3 = gen_docs.gen_supervision_note(
            order2, os.path.join(jd_dir, '监理情况说明.docx'))
        for p in (q1, q2, q3):
            assert os.path.exists(p) and os.path.getsize(p) > 0, f'文件未生成: {p}'
        print(f'[4] 四方交底全套 3 个文件生成：通过（{jd_dir}）')

        # ---- 5. 导出 Excel（按界面同样的方式组装行）----
        now = datetime(2024, 1, 8, 12, 0)  # 固定“当前时间”保证结果可复现
        rows = []
        for o in store.list_orders():
            start = datetime.strptime(o['start_time'], '%Y-%m-%d %H:%M:%S')
            v_dl, f_dl = worktime.compute_deadlines(o['flow_type'], start, holidays, extra)
            earlier = min(d for d in (v_dl, f_dl) if d is not None)
            rows.append(dict(
                o,
                visit_deadline=_fmt_dt(v_dl),
                full_deadline=_fmt_dt(f_dl),
                remaining=worktime.remaining_text(now, earlier),
            ))
        xlsx_path = os.path.join(out_dir, '工单列表.xlsx')
        export.export_orders(xlsx_path, rows)
        assert os.path.exists(xlsx_path) and os.path.getsize(xlsx_path) > 0
        # 回读校验
        from openpyxl import load_workbook
        ws = load_workbook(xlsx_path).active
        assert ws.max_row == 4, f'Excel 应为 1 表头 + 3 数据行，实际 {ws.max_row} 行'
        assert ws.cell(row=2, column=1).value == 'GD20240106001'
        # 新增两列：'地址'、'是否建群' 位于 '状态' 列之前
        assert ws.cell(row=1, column=10).value == '地址', f"表头第10列应为'地址'，实际 {ws.cell(row=1, column=10).value}"
        assert ws.cell(row=1, column=11).value == '是否建群', f"表头第11列应为'是否建群'，实际 {ws.cell(row=1, column=11).value}"
        assert ws.cell(row=1, column=12).value == '状态', f"表头第12列应为'状态'，实际 {ws.cell(row=1, column=12).value}"
        print(f'[5] 导出 Excel 并回读校验：通过（{xlsx_path}）')

        # ---- 6. 汇总统计 ----
        # 契约：工单3 置为 status='流程办结'、in_group='是'（不再使用旧 finish_order/'已办结'）
        finished = dict(store.get_order('GD20240108003'))
        finished['status'] = '流程办结'
        finished['in_group'] = '是'
        store.update_order('GD20240108003', finished)
        s = export.summary([dict(o) for o in store.list_orders()])
        assert s['total'] == 3
        assert s['by_type'] == {'低压居民新装': 1, '低压非居民新装': 1, '低压充电桩新装': 1}
        assert s['by_status'] == {'在办': 2, '流程办结': 1}
        print('[6] 汇总统计（按类型 / 按状态）：通过')

        # ---- 6b. '上门服务办结' 状态仍出现在导出/汇总中且状态原样统计 ----
        visit_done = dict(store.get_order('GD20240106001'))
        visit_done['status'] = '上门服务办结'
        visit_done['in_group'] = '是'
        store.update_order('GD20240106001', visit_done)
        s2 = export.summary([dict(o) for o in store.list_orders()])
        assert s2['total'] == 3, '上门服务办结工单仍应计入汇总总数'
        assert s2['by_status'] == {'上门服务办结': 1, '在办': 1, '流程办结': 1}, \
            f"'上门服务办结' 应原样统计，实际 {s2['by_status']}"
        export.export_orders(xlsx_path, [dict(o) for o in store.list_orders()])
        ws2 = load_workbook(xlsx_path).active
        assert ws2.max_row == 4, '上门服务办结工单仍应出现在导出中'
        statuses = [ws2.cell(row=r, column=12).value for r in range(2, 5)]
        assert '上门服务办结' in statuses, f'导出中应包含上门服务办结状态，实际 {statuses}'
        print("[6b] '上门服务办结' 工单导出/汇总可见且状态原样统计：通过")

        print('\n端到端集成冒烟全部通过。')
        return 0
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


if __name__ == '__main__':
    sys.exit(main())
