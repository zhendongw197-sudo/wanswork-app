# -*- coding: utf-8 -*-
"""供方单台账测试：store 供方单 API 与 export 供方单 Excel 导出。"""

import os
import shutil
import sys
import tempfile

# 保证可以 import app 包（项目根目录加入 sys.path）
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.export import SUPPLY_FIELDS, SUPPLY_HEADERS, export_supply_orders
from app.store import Store
from openpyxl import load_workbook


def _make_store():
    """在临时目录中建 Store，返回 (store, 临时目录)。"""
    d = tempfile.mkdtemp(prefix='workorder_supply_test_')
    return Store(d), d


def _sample_record(flow_id='GF20240108001'):
    """样例供方单记录：含生成字段、手工字段、materials、supply。"""
    return {
        'flow_id': flow_id,
        'manager': '李四',
        'archived': '否',
        'account_no': '3201240001234567',
        'account_name': '张三五金店',
        'address': '溧水区永阳街道xx路1号',
        'start_time': '2024-01-08 08:30:00',
        'push_time': '2024-01-09 10:00:00',
        'overdue_time': '2024-01-15 17:00:00',
        'progress_marketing': '已联系客户',
        'progress_equipment': '已发料',
        'remark': '首次录入备注',
        'contractor': '力诺建设',
        'work_desc': '新装 4*16 电缆 50 米',
        'materials': {'cable_4x16': 50, 'term_16': 2},
        'supply': {'voltage': '380', 'box_phase': '三相'},
    }


def test_supply_crud():
    """供方单增删改查。"""
    store, d = _make_store()
    try:
        store.add_supply_order(_sample_record())
        got = store.get_supply_order('GF20240108001')
        assert got is not None and got['account_name'] == '张三五金店'
        assert got['materials'] == {'cable_4x16': 50, 'term_16': 2}
        assert got['supply']['voltage'] == '380'
        assert len(store.list_supply_orders()) == 1

        updated = _sample_record()
        updated['account_name'] = '张三建材店'
        store.update_supply_order('GF20240108001', updated)
        assert store.get_supply_order('GF20240108001')['account_name'] == '张三建材店'

        store.delete_supply_order('GF20240108001')
        assert store.get_supply_order('GF20240108001') is None
        assert store.list_supply_orders() == []
    finally:
        shutil.rmtree(d, ignore_errors=True)


def test_supply_flow_id_unique():
    """flow_id 重复时 add_supply_order 抛 ValueError。"""
    store, d = _make_store()
    try:
        store.add_supply_order(_sample_record())
        try:
            store.add_supply_order(_sample_record())
        except ValueError:
            pass
        else:
            raise AssertionError('重复 flow_id 未抛 ValueError')
    finally:
        shutil.rmtree(d, ignore_errors=True)


def test_supply_update_missing_raises():
    """update_supply_order 更新不存在的记录抛 KeyError。"""
    store, d = _make_store()
    try:
        try:
            store.update_supply_order('GF999', _sample_record('GF999'))
        except KeyError:
            pass
        else:
            raise AssertionError('更新不存在的供方单未抛 KeyError')
    finally:
        shutil.rmtree(d, ignore_errors=True)


def test_supply_upsert_insert():
    """upsert 新记录：整条插入，手工字段缺省补 ''。"""
    store, d = _make_store()
    try:
        record = _sample_record()
        for key in ('archived', 'progress_marketing', 'progress_equipment',
                    'remark', 'contractor'):
            record.pop(key)
        store.upsert_supply_order(record)
        got = store.get_supply_order('GF20240108001')
        assert got is not None and got['account_name'] == '张三五金店'
        for key in ('archived', 'progress_marketing', 'progress_equipment',
                    'remark', 'contractor'):
            assert got[key] == '', f'手工字段 {key} 未补空串'
    finally:
        shutil.rmtree(d, ignore_errors=True)


def test_supply_upsert_keeps_manual_fields():
    """upsert 已存在记录：手工字段保留，生成字段被刷新。"""
    store, d = _make_store()
    try:
        store.add_supply_order(_sample_record())

        # 模拟界面手工编辑：改 5 个手工字段
        edited = store.get_supply_order('GF20240108001')
        edited['archived'] = '是'
        edited['progress_marketing'] = '用户已确认方案'
        edited['progress_equipment'] = '未发料'
        edited['remark'] = '手工备注'
        edited['contractor'] = '手工配套单位'
        store.update_supply_order('GF20240108001', edited)

        # 再次生成并 upsert：生成字段变化、手工字段带脏值（应被忽略）
        regenerated = _sample_record()
        regenerated['account_name'] = '张三五金店（新户名）'
        regenerated['address'] = '新地址 2 号'
        regenerated['work_desc'] = '新装 4*35 电缆 80 米'
        regenerated['materials'] = {'cable_4x35': 80, 'term_35': 2}
        regenerated['supply'] = {'voltage': '220', 'box_phase': '单相'}
        regenerated['remark'] = '这条不应生效'  # 手工字段脏值
        store.upsert_supply_order(regenerated)

        got = store.get_supply_order('GF20240108001')
        # 生成字段被刷新
        assert got['account_name'] == '张三五金店（新户名）'
        assert got['address'] == '新地址 2 号'
        assert got['work_desc'] == '新装 4*35 电缆 80 米'
        assert got['materials'] == {'cable_4x35': 80, 'term_35': 2}
        assert got['supply']['voltage'] == '220'
        # 5 个手工字段保留旧值
        assert got['archived'] == '是'
        assert got['progress_marketing'] == '用户已确认方案'
        assert got['progress_equipment'] == '未发料'
        assert got['remark'] == '手工备注'
        assert got['contractor'] == '手工配套单位'

        # 跨实例持久化
        store2 = Store(d)
        got2 = store2.get_supply_order('GF20240108001')
        assert got2['remark'] == '手工备注'
        assert got2['account_name'] == '张三五金店（新户名）'
    finally:
        shutil.rmtree(d, ignore_errors=True)


def test_supply_old_orders_unaffected():
    """供方单操作不影响原有工单存储。"""
    store, d = _make_store()
    try:
        store.add_order({'flow_id': 'GD1', 'status': '在办'})
        store.add_supply_order(_sample_record())
        assert store.get_order('GD1')['status'] == '在办'
        assert len(store.list_supply_orders()) == 1
        store.delete_supply_order('GF20240108001')
        assert store.get_order('GD1') is not None
    finally:
        shutil.rmtree(d, ignore_errors=True)


def test_export_supply_orders():
    """导出供方单台账：表头 31 列、材料值落在正确列。"""
    d = tempfile.mkdtemp(prefix='workorder_supply_test_')
    try:
        path = os.path.join(d, 'supply.xlsx')
        records = [_sample_record()]
        export_supply_orders(path, records)
        assert os.path.exists(path)

        wb = load_workbook(path)
        ws = wb['供方单台账']

        # 表头共 31 列且内容一致
        headers = [ws.cell(row=1, column=i).value for i in range(1, len(SUPPLY_HEADERS) + 1)]
        assert len(SUPPLY_HEADERS) == 31
        assert headers == SUPPLY_HEADERS
        assert ws.max_column == 31

        # 表头样式：加粗 + 填充 + 冻结首行
        assert ws.cell(row=1, column=1).font.bold
        assert ws.cell(row=1, column=1).fill.fgColor.rgb is not None
        assert ws.freeze_panes == 'A2'

        # 基本字段列：逐列按 SUPPLY_FIELDS 校验
        row = [ws.cell(row=2, column=i).value for i in range(1, 32)]
        assert row[0] == '李四'                       # 客户经理
        assert row[1] == '否'                         # 是否归档
        assert row[2] == '3201240001234567'           # 户号
        assert row[3] == 'GF20240108001'              # 流程编号
        assert row[4] == '张三五金店'                  # 户名
        assert row[13] == '新装 4*16 电缆 50 米'       # 现场勘察方案及工作量

        # 材料列：cable_4x16 在第 16 列（14 + 2），term_16 在第 26 列（14 + 12）
        idx_4x16 = SUPPLY_FIELDS.index('cable_4x16')
        idx_term16 = SUPPLY_FIELDS.index('term_16')
        assert row[idx_4x16] == '50'
        assert row[idx_term16] == '2'
        # 未填写的材料列为空串
        idx_overhead = SUPPLY_FIELDS.index('overhead')
        assert row[idx_overhead] in ('', None)
        # 表头与材料列对应
        assert ws.cell(row=1, column=idx_4x16 + 1).value == '4*16'
        assert ws.cell(row=1, column=idx_term16 + 1).value == '16冷缩终端'
        assert ws.cell(row=1, column=idx_overhead + 1).value == '架空导线'

        # materials 缺省时也能导出
        path2 = os.path.join(d, 'supply2.xlsx')
        export_supply_orders(path2, [{'flow_id': 'GF2', 'manager': '王五'}])
        ws2 = load_workbook(path2)['供方单台账']
        assert ws2.cell(row=2, column=1).value == '王五'
        assert ws2.cell(row=2, column=15).value in ('', None)
    finally:
        shutil.rmtree(d, ignore_errors=True)


if __name__ == '__main__':
    # pytest 不可用时可直接 python tests/test_store_supply.py 运行
    funcs = [(name, fn) for name, fn in sorted(globals().items())
             if name.startswith('test_') and callable(fn)]
    failed = 0
    for name, fn in funcs:
        try:
            fn()
            print(f'通过: {name}')
        except AssertionError as e:
            failed += 1
            print(f'失败: {name} -> {e}')
    print(f'\n共 {len(funcs)} 项，失败 {failed} 项')
    sys.exit(1 if failed else 0)
